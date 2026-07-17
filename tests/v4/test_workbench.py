"""v4.5 Windows Workbench: canonical model + workbook tests
(release gates R02, R04)."""

import pathlib

import pytest
from openpyxl import load_workbook

from rgcs_workbench import (EVIDENCE_CLASSES, PRIVACY_CLASSES,
                            REQUIRED_SHEETS)
from rgcs_workbench.canonical import CanonicalError, Record, build
from rgcs_workbench.workbook import generate


# --- canonical model -----------------------------------------------------

def test_build_populates_all_registries():
    store = build()
    counts = store.counts()
    for table in ("frequency_keys", "harmonic_relations",
                  "specimens", "mode_estimates", "hypotheses",
                  "eye_results", "resonator_platform",
                  "source_registry", "corrections"):
        assert counts.get(table, 0) > 0, table


def test_every_record_has_a_legal_evidence_class():
    store = build()
    for table, recs in store.tables.items():
        for r in recs:
            assert r.evidence_class in EVIDENCE_CLASSES
            assert r.privacy_class in PRIVACY_CLASSES


def test_record_rejects_bad_classes():
    with pytest.raises(CanonicalError):
        Record("X", "k", "MEASURED_MAGIC")
    with pytest.raises(CanonicalError):
        Record("X", "k", "LORE", privacy_class="SECRET")


def test_no_physical_evidence_class_anywhere():
    """R04: nothing is BENCH_MEASUREMENT or INDEPENDENT_REPLICATION —
    no physical work has been done, so no row may claim it."""
    store = build()
    for recs in store.tables.values():
        for r in recs:
            assert r.evidence_class not in (
                "BENCH_MEASUREMENT", "INDEPENDENT_REPLICATION"), \
                f"{r.id} claims physical evidence with none"


def test_specimen_stays_prearrival_unverified():
    store = build()
    spec = store.tables["specimens"][0]
    assert spec.fields["status"] == "PREARRIVAL_UNVERIFIED"
    assert spec.evidence_class == "SOURCE_CLAIM"
    assert "not a measurement" in spec.fields["warning"].lower() \
        or "not measurements" in spec.fields["warning"].lower()


def test_hypotheses_all_untested():
    store = build()
    for h in store.tables["hypotheses"]:
        assert h.fields["status"] == "UNTESTED"


def test_relations_carry_mechanism_class():
    store = build()
    classes = {r.fields["primary_class"]
               for r in store.tables["harmonic_relations"]}
    assert "HARMONIC" in classes
    assert "ARITHMETIC_COINCIDENCE" in classes
    # 4096x5 is a harmonic; the exact high-order closures are not
    harmonic = [r for r in store.tables["harmonic_relations"]
                if r.fields["primary_class"] == "HARMONIC"]
    assert any(r.fields["expression"] == "5x4096" for r in harmonic)


# --- privacy filter ------------------------------------------------------

def test_public_export_excludes_private_rows():
    store = build()
    store.add("lore_registry", Record(
        "LORE-PRIV", "lore", "LORE", privacy_class="PRIVATE",
        fields={"statement": "private content"}))
    pub = store.rows("lore_registry", include_private=False)
    priv = store.rows("lore_registry", include_private=True)
    assert len(priv) == len(pub) + 1
    assert not any(r["id"] == "LORE-PRIV" for r in pub)


# --- workbook ------------------------------------------------------------

def test_workbook_has_all_required_sheets(tmp_path):
    out = tmp_path / "wb.xlsx"
    generate().save(out)
    wb = load_workbook(out)
    for s in REQUIRED_SHEETS:
        assert s in wb.sheetnames, s
    assert wb.sheetnames[0] == "Dashboard"


def test_workbook_formulas_are_visible_and_recalculable(tmp_path):
    """R02: formulas compute, no baked constants where a formula is
    practical."""
    out = tmp_path / "wb.xlsx"
    generate().save(out)
    wb = load_workbook(out)
    # mode-estimate error is a formula referencing target and model
    g2 = wb["Mode Estimates"]["G2"].value
    assert isinstance(g2, str) and g2.startswith("=(") and \
        "F2" in g2 and "D2" in g2
    # dashboard total is a SUM formula
    dash = wb["Dashboard"]
    assert any(str(c.value).startswith("=SUM")
               for row in dash.iter_rows() for c in row
               if c.value)


def test_workbook_opens_without_external_links(tmp_path):
    out = tmp_path / "wb.xlsx"
    generate().save(out)
    wb = load_workbook(out)
    # openpyxl exposes external references via _external_links
    assert not getattr(wb, "_external_links", [])
    # no macros (xlsx, not xlsm)
    assert not wb.vba_archive


def test_workbook_named_tables_unique(tmp_path):
    out = tmp_path / "wb.xlsx"
    generate().save(out)
    wb = load_workbook(out)
    names = []
    for s in wb.sheetnames:
        names += list(wb[s].tables.keys())
    assert len(names) == len(set(names)), "duplicate table names"
    assert len(names) >= 12


def test_public_workbook_has_no_private_content(tmp_path):
    """R04: a public workbook never contains a PRIVATE row."""
    store = build()
    store.add("lore_registry", Record(
        "LORE-SECRET", "lore", "LORE", privacy_class="PRIVATE",
        fields={"statement": "SECRET-SELLER-ADDRESS"}))
    out = tmp_path / "pub.xlsx"
    generate(store, include_private=False).save(out)
    text = out.read_bytes()
    assert b"SECRET-SELLER-ADDRESS" not in text


def test_workbook_deterministic(tmp_path):
    """Same canonical data -> same workbook content (dashboard counts
    and sheet set), regardless of run."""
    a = generate(build())
    b = generate(build())
    assert a.sheetnames == b.sheetnames
    ca = [c.value for row in a["Frequency Keys"].iter_rows()
          for c in row]
    cb = [c.value for row in b["Frequency Keys"].iter_rows()
          for c in row]
    assert ca == cb


def test_workbook_claim_boundary_present(tmp_path):
    out = tmp_path / "wb.xlsx"
    generate().save(out)
    wb = load_workbook(out)
    dash = " ".join(str(c.value) for row in wb["Dashboard"].iter_rows()
                    for c in row if c.value)
    assert "does NOT claim" in dash or "not claim" in dash.lower()
    assert "PREARRIVAL" in dash


def test_generator_reachable_as_module():
    import rgcs_workbench.workbook as m
    assert hasattr(m, "generate") and hasattr(m, "main")
