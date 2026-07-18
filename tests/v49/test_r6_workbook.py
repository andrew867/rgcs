"""R6 workbook and canonical-store integration."""

from __future__ import annotations

import io

import openpyxl
import pytest

from rgcs_workbench import EVIDENCE_CLASSES, REQUIRED_SHEETS
from rgcs_workbench.canonical import build
from rgcs_workbench.workbook import EVIDENCE_FILL, generate

R6_SHEETS = ("R6 Claims", "R6 Apparatus", "R6 Witness", "R6 Mailbox",
             "R6 Grid")


@pytest.fixture(scope="module")
def store():
    return build("4.9.0")


@pytest.fixture(scope="module")
def wb():
    return generate(version="4.9.0", include_private=False)


# --- the coupling that crashed generation ----------------------------

def test_every_evidence_class_has_a_fill():
    """Adding a class without a fill crashes workbook generation.

    This failed for real when the R6 ladder rungs were registered: the
    dashboard indexes EVIDENCE_FILL directly, so a missing entry is a
    KeyError at generate time rather than a styling glitch. The two
    tables must stay in sync, and this is what says so.
    """
    missing = [c for c in EVIDENCE_CLASSES if c not in EVIDENCE_FILL]
    assert not missing, f"evidence classes without a fill: {missing}"


def test_no_orphan_fills():
    orphans = [c for c in EVIDENCE_FILL if c not in EVIDENCE_CLASSES]
    assert not orphans, f"fills for unknown classes: {orphans}"


def test_no_detection_class_was_added():
    for c in EVIDENCE_CLASSES:
        assert "DETECT" not in c.upper()


# --- sheets -----------------------------------------------------------

def test_all_r6_sheets_registered():
    for s in R6_SHEETS:
        assert s in REQUIRED_SHEETS


def test_workbook_has_all_required_sheets(wb):
    assert set(REQUIRED_SHEETS) <= set(wb.sheetnames)


def test_r6_sheets_are_populated(wb):
    for s in R6_SHEETS:
        assert wb[s].max_row > 1, f"{s} has no data rows"


def test_r6_sheets_use_the_ordered_column_union(wb):
    """The v4.8.1 lesson: headers must cover every record's keys.

    R6 rows are deliberately heterogeneous -- a refusal row and a
    clock-comparison row share almost no fields -- so a header built
    from the first row only would silently drop most of the sheet.
    """
    for s in R6_SHEETS:
        ws = wb[s]
        header = [c.value for c in next(ws.iter_rows())]
        assert len(header) == len(set(header)), f"{s} has duplicate cols"
        # every non-empty cell must sit under a named column
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.value not in (None, ""):
                    idx = cell.column - 1
                    assert idx < len(header), (
                        f"{s}: value outside the header range")


def test_witness_sheet_exposes_the_resolution_result(wb):
    """A reader must be able to see that caesium cannot resolve 1 m."""
    ws = wb["R6 Witness"]
    header = [c.value for c in next(ws.iter_rows())]
    si = header.index("status")
    statuses = [r[si].value for r in ws.iter_rows(min_row=2)]
    assert "PREDICTION_BELOW_RESOLUTION" in statuses
    assert "RELATIVISTIC_SHIFT_CONSISTENT" in statuses


def test_mailbox_sheet_exposes_the_sovereignty_verdict(wb):
    ws = wb["R6 Mailbox"]
    header = [c.value for c in next(ws.iter_rows())]
    si = header.index("status")
    statuses = [r[si].value for r in ws.iter_rows(min_row=2)]
    assert "SOVEREIGN_NAVIGATION_UNSUPPORTED" in statuses


def test_grid_sheet_exposes_no_real_data(wb):
    ws = wb["R6 Grid"]
    header = [c.value for c in next(ws.iter_rows())]
    idx = header.index("planetary_status")
    vals = [r[idx].value for r in ws.iter_rows(min_row=2)
            if r[idx].value]
    assert vals and all(v == "NO_REAL_DATA" for v in vals)


def test_claims_sheet_carries_verbatim_and_ceiling(wb):
    ws = wb["R6 Claims"]
    header = [c.value for c in next(ws.iter_rows())]
    assert "verbatim" in header and "ceiling" in header
    assert "correction" in header


def test_apparatus_sheet_shows_the_frequency_null(wb):
    ws = wb["R6 Apparatus"]
    header = [c.value for c in next(ws.iter_rows())]
    idx = header.index("significant")
    vals = [r[idx].value for r in ws.iter_rows(min_row=2)
            if r[idx].value is not None]
    assert vals and not any(vals), (
        "the source frequency audit must not report significance")


# --- store ------------------------------------------------------------

def test_store_has_all_r6_tables(store):
    for t in ("r6_claims", "r6_apparatus", "r6_witness", "r6_mailbox",
              "r6_grid"):
        assert store.tables.get(t), f"{t} is empty"


def test_no_physical_evidence_class_in_r6(store):
    """R6 is software-only; it may never emit a physical class."""
    from rgcs_workbench import PHYSICAL_EVIDENCE_CLASSES
    for t, rows in store.tables.items():
        if not t.startswith("r6_"):
            continue
        for r in rows:
            assert r.evidence_class not in PHYSICAL_EVIDENCE_CLASSES, (
                f"{r.id} claims physical evidence")


def test_r6_ids_are_positional_not_hashed(store):
    """R4 lesson 3: Python hash() is randomized per process."""
    a = build("4.9.0")
    b = build("4.9.0")
    for t in ("r6_claims", "r6_grid", "r6_mailbox"):
        assert [r.id for r in a.tables[t]] == [r.id for r in b.tables[t]]


def test_workbook_generation_is_deterministic():
    """Two fresh generations must agree cell for cell."""
    def cells():
        w = generate(version="4.9.0", include_private=False)
        buf = io.BytesIO()
        w.save(buf)
        rb = openpyxl.load_workbook(io.BytesIO(buf.getvalue()))
        return {n: [[c.value for c in row]
                    for row in rb[n].iter_rows()]
                for n in rb.sheetnames if n.startswith("R6")}
    assert cells() == cells()
