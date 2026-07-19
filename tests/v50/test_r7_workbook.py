"""R7 workbook and canonical-store integration."""

from __future__ import annotations

import io

import openpyxl
import pytest

from rgcs_workbench import EVIDENCE_CLASSES, REQUIRED_SHEETS
from rgcs_workbench.canonical import build
from rgcs_workbench.workbook import EVIDENCE_FILL, generate

R7_SHEETS = ("R7 CW", "R7 Gravity", "R7 Clock Link", "R7 Governance")


@pytest.fixture(scope="module")
def store():
    return build("5.0.0")


@pytest.fixture(scope="module")
def wb():
    return generate(version="5.0.0", include_private=False)


def test_evidence_classes_and_fills_stay_in_sync():
    """The R6-D-006 regression guard, still holding."""
    assert not [c for c in EVIDENCE_CLASSES if c not in EVIDENCE_FILL]
    assert not [c for c in EVIDENCE_FILL if c not in EVIDENCE_CLASSES]


def test_all_r7_sheets_registered():
    for s in R7_SHEETS:
        assert s in REQUIRED_SHEETS


def test_r7_sheets_are_populated(wb):
    for s in R7_SHEETS:
        assert wb[s].max_row > 1, f"{s} has no data rows"


def test_r7_sheets_use_the_ordered_column_union(wb):
    for s in R7_SHEETS:
        ws = wb[s]
        header = [c.value for c in next(ws.iter_rows())]
        assert len(header) == len(set(header))
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.value not in (None, ""):
                    assert cell.column - 1 < len(header)


def test_cw_sheet_shows_zero_informative_bits(wb):
    """A reader must be able to see the headline was refuted."""
    ws = wb["R7 CW"]
    header = [c.value for c in next(ws.iter_rows())]
    idx = header.index("informative_bits")
    vals = [r[idx].value for r in ws.iter_rows(min_row=2)
            if r[idx].value is not None]
    assert vals and all(v == 0 for v in vals)


def test_cw_sheet_shows_the_null_probability(wb):
    ws = wb["R7 CW"]
    header = [c.value for c in next(ws.iter_rows())]
    idx = header.index("null_3x12_p_both")
    vals = [r[idx].value for r in ws.iter_rows(min_row=2)
            if r[idx].value is not None]
    assert vals and all(v == 1.0 for v in vals)


def test_gravity_sheet_is_all_refused(wb):
    ws = wb["R7 Gravity"]
    header = [c.value for c in next(ws.iter_rows())]
    idx = header.index("status")
    vals = [r[idx].value for r in ws.iter_rows(min_row=2)
            if r[idx].value]
    assert vals and all(v == "REFUSED_BY_ARITHMETIC" for v in vals)


def test_clock_sheet_shows_quartz_unresolvable(wb):
    ws = wb["R7 Clock Link"]
    header = [c.value for c in next(ws.iter_rows())]
    oi, ri = header.index("oscillator"), header.index("resolvable")
    rows = [(r[oi].value, r[ri].value) for r in ws.iter_rows(min_row=2)
            if r[oi].value]
    ocxo = [res for osc, res in rows if osc == "OCXO"]
    assert ocxo and not any(ocxo)


def test_clock_sheet_shows_at_least_one_resolvable_pairing(wb):
    """Optical on stabilized fibre must be visible as achievable."""
    ws = wb["R7 Clock Link"]
    header = [c.value for c in next(ws.iter_rows())]
    ri = header.index("resolvable")
    vals = [r[ri].value for r in ws.iter_rows(min_row=2)
            if r[ri].value is not None]
    assert any(vals)


def test_governance_sheet_shows_private_rc_not_authorized(wb):
    ws = wb["R7 Governance"]
    header = [c.value for c in next(ws.iter_rows())]
    idx = header.index("authorized_to_publish")
    vals = [r[idx].value for r in ws.iter_rows(min_row=2)
            if r[idx].value is not None]
    assert vals and not any(vals)


def test_governance_sheet_records_the_path(wb):
    ws = wb["R7 Governance"]
    header = [c.value for c in next(ws.iter_rows())]
    idx = header.index("path")
    vals = [r[idx].value for r in ws.iter_rows(min_row=2)
            if r[idx].value]
    assert "PRIVATE_RC" in vals


def test_store_has_all_r7_tables(store):
    for t in ("r7_cw", "r7_gravity", "r7_clocklink", "r7_governance"):
        assert store.tables.get(t), f"{t} is empty"


def test_no_physical_evidence_class_in_r7(store):
    from rgcs_workbench import PHYSICAL_EVIDENCE_CLASSES
    for t, rows in store.tables.items():
        if not t.startswith("r7_"):
            continue
        for r in rows:
            assert r.evidence_class not in PHYSICAL_EVIDENCE_CLASSES


def test_r7_ids_are_deterministic():
    a, b = build("5.0.0"), build("5.0.0")
    for t in ("r7_cw", "r7_gravity", "r7_clocklink"):
        assert [r.id for r in a.tables[t]] == [r.id for r in b.tables[t]]


def test_workbook_generation_is_deterministic():
    def cells():
        w = generate(version="5.0.0", include_private=False)
        buf = io.BytesIO()
        w.save(buf)
        rb = openpyxl.load_workbook(io.BytesIO(buf.getvalue()))
        return {n: [[c.value for c in row] for row in rb[n].iter_rows()]
                for n in rb.sheetnames if n.startswith("R7")}
    assert cells() == cells()
