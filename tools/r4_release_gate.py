"""R04/R65 — the no-post-tag-sync release gate.

v4.6, v4.7.0 and v4.7.1 each needed a commit AFTER the tag to sync the
committed workbook and checksum manifest with what was actually
published. That means the tagged tree was never quite the released
tree. R4 forbids it: the tag must already contain every release-owned
artifact.

This gate is run BEFORE tagging and refuses if any committed
release-owned artifact is stale:

- ``release/v45/RGCS_Master_Evidence_Workbook.xlsx`` must match a
  freshly generated public workbook (content-wise);
- ``release/v45/SHA256SUMS.txt`` must list that workbook's hash.

**Two artifacts are structurally impossible to pre-commit, and saying
so precisely is part of the gate.**

1. *The source archive* is ``git archive`` of the tagged commit, so it
   CONTAINS the manifest. A committed manifest listing the source
   archive's own hash is circular.
2. *The frozen Windows binaries* embed the tagged commit id in
   ``_build_stamp.json`` (the v4.5.2 anti-stale mechanism), so they can
   only be built AFTER the commit exists. A committed manifest listing
   their hashes would have to predict the hash of a file that contains
   that very commit's id.

So the COMMITTED manifest covers the workbook — genuinely
pre-committable, since the workbook is generated from source and
embeds no commit id. The COMPLETE manifest (standard bundle + Windows
assets + workbook) is generated at publish time and uploaded as a
release asset. An upload is not repository content, so it requires no
commit and leaves the tag complete in the only sense that is
achievable.

    python tools/r4_release_gate.py            # verify (exit 1 on stale)
    python tools/r4_release_gate.py --write    # refresh, then verify
"""

from __future__ import annotations

import hashlib
import io
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

V45 = ROOT / "release" / "v45"
WORKBOOK = V45 / "RGCS_Master_Evidence_Workbook.xlsx"
MANIFEST = V45 / "SHA256SUMS.txt"

#: Artifacts the tag must already contain, with the glob that finds the
#: built Windows assets for the current version.
WINDOWS_GLOBS = ("RGCS-Workbench-*-Windows-x64-portable.zip",
                 "RGCS-Workbench-*-Windows-x64-Setup.exe")


def _version() -> str:
    import re
    m = re.search(r'^version = "([^"]+)"',
                  (ROOT / "pyproject.toml").read_text(encoding="utf-8"),
                  re.M)
    return m.group(1) if m else "0.0.0"


def generate_workbook_bytes(version: str) -> bytes:
    """A freshly generated PUBLIC workbook, as bytes."""
    from rgcs_workbench.workbook import generate
    wb = generate(version=version, include_private=False)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sha(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()


def workbook_is_current(version: str) -> dict:
    """openpyxl embeds no timestamp we control, but zip member order and
    content are deterministic for a deterministic store, so a byte
    comparison is meaningful. Content equality is checked cell-wise as
    the authoritative test, with bytes reported for information."""
    if not WORKBOOK.exists():
        return {"ok": False, "reason": "committed workbook missing"}
    fresh = generate_workbook_bytes(version)
    on_disk = WORKBOOK.read_bytes()
    if _sha(fresh) == _sha(on_disk):
        return {"ok": True, "match": "bytes"}
    # fall back to cell-wise comparison (zip metadata can differ)
    import openpyxl
    a = openpyxl.load_workbook(io.BytesIO(fresh))
    b = openpyxl.load_workbook(io.BytesIO(on_disk))
    if a.sheetnames != b.sheetnames:
        return {"ok": False, "reason": "sheet set differs",
                "fresh_sheets": len(a.sheetnames),
                "committed_sheets": len(b.sheetnames)}
    for name in a.sheetnames:
        va = [[c.value for c in row] for row in a[name].iter_rows()]
        vb = [[c.value for c in row] for row in b[name].iter_rows()]
        if va != vb:
            return {"ok": False,
                    "reason": f"sheet {name!r} content differs"}
    return {"ok": True, "match": "cell-wise"}


def windows_assets(version: str) -> list:
    out = []
    for g in WINDOWS_GLOBS:
        out.extend(sorted(V45.glob(g.replace("*", version))))
    return out


def manifest_is_current(version: str) -> dict:
    if not MANIFEST.exists():
        return {"ok": False, "reason": "committed manifest missing"}
    listed = {}
    for line in MANIFEST.read_text(encoding="utf-8").splitlines():
        if "  " in line:
            h, n = line.split("  ", 1)
            listed[n.strip()] = h
    # Only the workbook is pre-committable (see module docstring):
    # the frozen binaries embed the tagged commit id, so their hashes
    # cannot exist before that commit does.
    required = [WORKBOOK]
    problems = []
    for p in required:
        if p.name not in listed:
            problems.append(f"{p.name}: not listed")
        elif listed[p.name] != _sha(p.read_bytes()):
            problems.append(f"{p.name}: hash stale")
    return {"ok": not problems, "problems": problems,
            "listed": len(listed),
            "note": "Windows-asset hashes are intentionally absent: "
                    "they embed the tagged commit and are covered by "
                    "the uploaded manifest"}


def write_release_owned(version: str) -> dict:
    """Refresh the committed artifacts so the tag can contain them."""
    V45.mkdir(parents=True, exist_ok=True)
    WORKBOOK.write_bytes(generate_workbook_bytes(version))
    lines = []
    for p in [WORKBOOK]:
        lines.append(f"{_sha(p.read_bytes())}  {p.name}")
    MANIFEST.write_text("\n".join(lines) + "\n", encoding="utf-8",
                        newline="\n")
    return {"workbook": str(WORKBOOK.relative_to(ROOT)),
            "manifest_entries": len(lines)}


def gate(version: str | None = None) -> dict:
    v = version or _version()
    wb = workbook_is_current(v)
    mf = manifest_is_current(v)
    ok = wb["ok"] and mf["ok"]
    return {
        "version": v,
        "workbook": wb,
        "manifest": mf,
        "ok": ok,
        "verdict": "TAG_MAY_PROCEED" if ok else "REFUSE_TAG",
        "rule": "the tag must already contain every release-owned "
                "artifact; no post-tag synchronization commit is "
                "permitted (R04/R65)",
    }


def main() -> int:
    v = _version()
    if "--write" in sys.argv:
        print("refreshed:", write_release_owned(v))
    rep = gate(v)
    print(f"release gate for v{rep['version']}: {rep['verdict']}")
    print(f"  workbook: {rep['workbook']}")
    print(f"  manifest: {rep['manifest']}")
    return 0 if rep["ok"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
