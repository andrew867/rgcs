"""Master Evidence Workbook generator (v4.5 pack, P01/P05,
04_WORKBOOK_CONTRACT).

canonical store -> .xlsx, one-way. The workbook opens with no
external links, no macros, visible+recalculable formulas, stable
sheet names, named tables, frozen headers, conditional formatting for
evidence status, and a dashboard. Derived values are FORMULAS where
practical (so a reader can audit them), not baked constants.

Two export modes: PUBLIC (excludes PRIVATE rows) and PRIVATE (local,
may include private rows). Neither ever contains credentials or
private seller/order details — those are PRIVATE privacy_class and
filtered, and the specimen listing id is PUBLIC_SAFE (the item
number) without seller contact data."""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from . import EVIDENCE_CLASSES, REQUIRED_SHEETS, SCHEMA_VERSION
from .canonical import CanonicalStore, build

# evidence-class colours (weakest -> strongest); used in conditional
# formatting so a reader sees the class at a glance
EVIDENCE_FILL = {
    "LORE": "E0D0F0", "SOURCE_CLAIM": "FFF0C0",
    "DERIVED_ARITHMETIC": "D8E8FF", "ANALYTIC_MODEL": "CFE8CF",
    "NUMERICAL_SIMULATION": "BFE0D8", "SYNTHETIC_RUN": "FFE0B0",
    "BENCH_MEASUREMENT": "B0F0B0", "INDEPENDENT_REPLICATION":
    "80E080",
    # v4.6: terminal statuses, not ladder rungs. Red-grey so an
    # UNSUPPORTED row is visibly not an achievement.
    "UNSUPPORTED": "F0C8C8", "NOT_APPLICABLE": "D8D8D8",
}
HEADER_FILL = PatternFill("solid", fgColor="2E3A46")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _write_table(ws, rows: list, table_name: str,
                 start_row: int = 1) -> int:
    """Write a list of dicts as a named table with a styled header,
    frozen panes, and per-row evidence-class shading. Returns the last
    row index."""
    if not rows:
        ws.cell(start_row, 1, "(no rows)")
        return start_row
    cols = list(rows[0].keys())
    for j, c in enumerate(cols, 1):
        cell = ws.cell(start_row, j, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left")
    ev_col = cols.index("evidence_class") + 1 if "evidence_class" \
        in cols else None
    for i, row in enumerate(rows, start_row + 1):
        for j, c in enumerate(cols, 1):
            ws.cell(i, j, row.get(c))
        if ev_col:
            ev = row.get("evidence_class")
            fill = EVIDENCE_FILL.get(ev)
            if fill:
                for j in range(1, len(cols) + 1):
                    ws.cell(i, j).fill = PatternFill("solid",
                                                     fgColor=fill)
    last = start_row + len(rows)
    ref = f"A{start_row}:{get_column_letter(len(cols))}{last}"
    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight9", showRowStripes=False)
    ws.add_table(tbl)
    ws.freeze_panes = f"A{start_row + 1}"
    for j, c in enumerate(cols, 1):
        width = max(len(str(c)),
                    *(len(str(r.get(c, ""))) for r in rows[:50]))
        ws.column_dimensions[get_column_letter(j)].width = \
            min(max(width + 2, 10), 60)
    return last


def _dashboard(ws, store: CanonicalStore, version: str) -> None:
    ws["A1"] = "RGCS Master Evidence Workbook"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = (f"Schema {SCHEMA_VERSION} · RGCS v{version} · "
                "generated from canonical repository data")
    ws["A3"] = ("Lore proposes. Mathematics translates. Software "
                "attacks. Evidence decides. Provenance remembers.")
    ws["A3"].font = Font(italic=True)
    ws["A5"] = ("CLAIM BOUNDARY: this workbook shows arithmetic, "
                "analytic models, simulations, and synthetic runs. "
                "It does NOT claim any physical crystal resonance, "
                "the Eye hypothesis, or any anomalous channel is "
                "validated. The purchased specimen is PREARRIVAL, "
                "unmeasured.")
    ws["A5"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A5:H7")
    # counts table with a live COUNTA-based total formula
    ws["A9"] = "Registry"
    ws["B9"] = "Rows"
    ws["A9"].font = ws["B9"].font = Font(bold=True)
    counts = store.counts()
    r = 10
    for table, _ in sorted(counts.items()):
        ws.cell(r, 1, table)
        ws.cell(r, 2, len(store.rows(table, include_private=True)))
        r += 1
    ws.cell(r, 1, "TOTAL")
    ws.cell(r, 1).font = Font(bold=True)
    ws.cell(r, 2, f"=SUM(B10:B{r - 1})")   # visible formula
    ws.cell(r, 2).font = Font(bold=True)
    # evidence-class legend
    ws.cell(r + 2, 1, "Evidence classes (weakest -> strongest):")
    ws.cell(r + 2, 1).font = Font(bold=True)
    for k, ev in enumerate(EVIDENCE_CLASSES):
        cell = ws.cell(r + 3 + k, 1, ev)
        cell.fill = PatternFill("solid",
                                fgColor=EVIDENCE_FILL[ev])
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 12


def _mode_estimates_with_formulas(ws, rows: list) -> None:
    """Mode-estimate sheet keeps the target error as a VISIBLE
    formula (contract: derived values remain formulas where
    practical): rel_error = (target - model)/target."""
    if not rows:
        ws["A1"] = "(no rows)"
        return
    cols = ["id", "evidence_class", "model_level", "quarter_wave_hz",
            "half_wave_hz", "target_hz", "rel_error_formula",
            "band_lo_hz", "band_hi_hz", "note"]
    for j, c in enumerate(cols, 1):
        cell = ws.cell(1, j, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, row in enumerate(rows, 2):
        ws.cell(i, 1, row["id"])
        ws.cell(i, 2, row["evidence_class"])
        ws.cell(i, 3, row["model_level"])
        ws.cell(i, 4, row["quarter_wave_hz"])
        ws.cell(i, 5, row["half_wave_hz"])
        ws.cell(i, 6, row["target_hz"])
        # the FORMULA a reader can audit and recalculate:
        ws.cell(i, 7, f"=(F{i}-D{i})/F{i}")
        ws.cell(i, 8, row["band_lo_hz"])
        ws.cell(i, 9, row["band_hi_hz"])
        ws.cell(i, 10, row["note"])
        fill = EVIDENCE_FILL.get(row["evidence_class"])
        if fill:
            for j in range(1, len(cols) + 1):
                ws.cell(i, j).fill = PatternFill("solid",
                                                 fgColor=fill)
    ws.freeze_panes = "A2"
    for j in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 16
    ws.column_dimensions["J"].width = 50


def _guide(ws) -> None:
    ws["A1"] = "Workbook Guide"
    ws["A1"].font = Font(bold=True, size=14)
    lines = [
        "",
        "This workbook is generated ONE-WAY from the RGCS canonical "
        "data store. It is a formula-visible summary and an "
        "interoperability export — it is NOT the database.",
        "",
        "Sheets:",
        "  Dashboard        registry counts (SUM formula) + evidence "
        "legend + claim boundary",
        "  Frequency Keys   the seed key registry with status/tags",
        "  Harmonic Relations  every relation with its mechanism "
        "class (a 4096x5 harmonic is not an 8x2560 closure)",
        "  Specimens        the PREARRIVAL eBay listing (seller "
        "geometry, UNMEASURED)",
        "  Mode Estimates   1-D screening with the target error as a "
        "LIVE FORMULA",
        "  Timing Recipes   drive families incl. controls",
        "  Hypotheses       6 preregistered, all UNTESTED",
        "  Evidence Ledger  every row's evidence + privacy class",
        "  Eye Results      canonical v4.1 + v4.2.1 records "
        "(computational, idealized)",
        "  Resonator Platform  the synthetic design-to-certificate "
        "campaign",
        "  Hardware BOM     Gen-0 parts (NOT purchased) + CYD board "
        "profiles (candidate)",
        "  Experiment Queue  8 campaigns, all hardware-required",
        "  Corrections      preserved correction records",
        "  Source Registry  5 papers with forbidden-transfer rules",
        "  Lore Registry    the lore POLICY only (private content is "
        "local-only)",
        "  Installer Metadata  release facts (unsigned, no telemetry)",
        "",
        "Evidence classes never hide. Nothing physical is validated. "
        "The specimen is unmeasured. The Eye is a computation.",
        "",
        "Regenerate: python -m rgcs_workbench.workbook <out.xlsx>",
    ]
    for i, ln in enumerate(lines, 2):
        ws.cell(i, 1, ln)
    ws.column_dimensions["A"].width = 100


def generate(store: CanonicalStore | None = None,
             version: str = "4.5.0",
             include_private: bool = False) -> Workbook:
    store = store or build(version)
    wb = Workbook()
    wb.remove(wb.active)

    table_map = {
        "Frequency Keys": ("frequency_keys", "FrequencyKeys"),
        "Harmonic Relations": ("harmonic_relations",
                               "HarmonicRelations"),
        "Specimens": ("specimens", "Specimens"),
        "Timing Recipes": ("timing_recipes", "TimingRecipes"),
        "Hypotheses": ("hypotheses", "Hypotheses"),
        "Eye Results": ("eye_results", "EyeResults"),
        "Resonator Platform": ("resonator_platform",
                               "ResonatorPlatform"),
        "Hardware BOM": ("hardware_bom", "HardwareBOM"),
        "Experiment Queue": ("experiment_queue", "ExperimentQueue"),
        "Corrections": ("corrections", "Corrections"),
        "Source Registry": ("source_registry", "SourceRegistry"),
        "Lore Registry": ("lore_registry", "LoreRegistry"),
        "Installer Metadata": ("installer_metadata",
                               "InstallerMetadata"),
        "CSCP Candidates": ("cspc_candidates", "CSCPCandidates"),
        "CSCP Tetrahedron": ("cspc_tetrahedron", "CSCPTetrahedron"),
        "CSCP Spacetime": ("cspc_spacetime", "CSCPSpacetime"),
        "CSCP Experiments": ("cspc_experiments", "CSCPExperiments"),
    }

    for sheet in REQUIRED_SHEETS:
        ws = wb.create_sheet(sheet)
        if sheet == "Dashboard":
            _dashboard(ws, store, version)
        elif sheet == "Mode Estimates":
            _mode_estimates_with_formulas(
                ws, store.rows("mode_estimates", include_private))
        elif sheet == "Evidence Ledger":
            ledger = []
            for table in store.tables:
                for row in store.rows(table, include_private):
                    ledger.append({
                        "table": table, "id": row["id"],
                        "kind": row["kind"],
                        "evidence_class": row["evidence_class"],
                        "privacy_class": row["privacy_class"],
                        "provenance": row["provenance"]})
            _write_table(ws, ledger, "EvidenceLedger")
        elif sheet == "Workbook Guide":
            _guide(ws)
        elif sheet in table_map:
            table, name = table_map[sheet]
            _write_table(ws, store.rows(table, include_private), name)
        else:
            ws["A1"] = "(sheet reserved)"
    return wb


def main() -> int:
    import sys
    out = sys.argv[1] if len(sys.argv) > 1 else \
        "RGCS_Master_Evidence_Workbook.xlsx"
    private = "--private" in sys.argv
    wb = generate(include_private=private)
    wb.save(out)
    print(f"workbook written: {out} "
          f"({'PRIVATE' if private else 'PUBLIC'}); "
          f"{len(wb.sheetnames)} sheets")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
